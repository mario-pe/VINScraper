import openpyxl
from prompt_toolkit.mouse_events import MouseButton
from selenium import webdriver
import undetected_chromedriver as uc
from selenium.common import ElementNotInteractableException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep

from config import username, password, FILE_PATH, REMOVE_CONTENT_URL, DRIVER_PATH
from locators import (
    google_approval_button,
    login_button,
    next_username_input_name,
    next_username_xpath,
    password_input_name,
    password_next_button,
    button_new_block_request_xpath,
    input_block_address_xpath,
    button_send_block,
    button_send_block_request,
    all_results_xpath,
    page_exists_block_address_xpath,
    image_tab_xpath,
    google_image_link_link_input_id,
    google_image_link_radio_button_id, images_tab_xpath, request_send_correctly, google_image_link_radio_button_label,
    image_url_input, button_next_block_request_xpath, page_exists_block_address_span, passed_address_span,
)

BLUE_VIN = "2C3CDZBT5LH210515"


PAGES_LINKS = []
IMAGES_LINKS = []

# VINS = [
#     "1hd1krp14kb602715",
#     "1hd1ydj16lb065180",
#     "1hd1ktc14jb673480",
#     "1hd1bxv14eb034736",
#     "1hd1kef14mb640278",
#     "1hd1jnv10eb023434",
# ]

#16.11.2023
VINS = [
    # 'WAUC4CF57JA057553',
    # 'WA1BNAFY7J2209278',
    # '4FMUS18168R001926',
    # 'SAJBD4BV4HCY40534',
    # '1C6SRFHT3MN563477',
    # '1C6SRFHT6NN111766',
    # '3CZRU6H16MM748096',
    # 'WDCTG4EB8JJ417348',
    # 'YV4A22PL4K1480175',
    # 'YV4AC2HKXM2513227',
    # 'WBXJG9C03N5V49513',
    'WAUTNAF57JA025820',
    'WF0DP3TH1G4114251',
    'W1N0G6EB9LF803456',
    '1C6SRFU95MN819712',
]
def load_vins():
    workbook = openpyxl.load_workbook(FILE_PATH)
    sheet = workbook.active
    for i in range(1, sheet.max_row + 1):
        VINS.append(sheet.cell(row=i, column=1).value.upper())


def init_driver():
    # import ipdb;ipdb.set_trace()
    return uc.Chrome(executable_path=DRIVER_PATH)

def login(driver):
    try:
        driver.get("https://www.google.com")
        WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, google_approval_button)).click()
        WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, login_button)).click()
        user_name_input_element = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.NAME, next_username_input_name))
        user_name_input_element.send_keys(username)
        driver.find_element(By.XPATH, next_username_xpath).click()
        password_input_name_element = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.NAME, password_input_name))
        password_input_name_element.send_keys(password)
        WebDriverWait(driver, 10).until(lambda d: d.find_element(By.ID, password_next_button)).click()
    except Exception as e:
        sleep(1)
        password_input_name_element = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.NAME, password_input_name))
        password_input_name_element.send_keys(password)
        WebDriverWait(driver, 10).until(lambda d: d.find_element(By.ID, password_next_button)).click()


def first_search(driver):
    search_textarea = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.TAG_NAME, "textarea"))
    search_textarea.send_keys("Dodge Challenger")
    search_textarea.submit()


def back_to_all_results(driver):
    driver.execute_script("window.scrollTo(0, 0);")
    sleep(0.5)
    driver.find_element(By.XPATH, all_results_xpath).click()


def new_search(driver, vin):
    sleep(1)
    driver.get("https://www.google.com")
    try:
        search_textarea = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.TAG_NAME, "textarea"))
    except Exception:
        import ipdb; ipdb.set_trace()
        driver.get("https://www.google.com")
        sleep(2)
        search_textarea = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.TAG_NAME, "textarea"))
    search_textarea.send_keys(vin)
    search_textarea.submit()


def links_scrapper(driver, vin):
    links = []
    search_textarea = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.TAG_NAME, "textarea"))
    search_textarea.clear()
    search_textarea.send_keys(vin)
    search_textarea.submit()
    las = driver.find_elements(By.TAG_NAME, "a")
    for index, la in enumerate(las):
        href = la.get_attribute("href")

        if href:
            if "translate.google" in href:
                continue
            upper_href = href.upper()
            if "bid.cars".upper() in upper_href and f"-{vin.upper()}" in upper_href:
                links.append(href)
            if "ucars.pro".upper() in upper_href and f"-{vin.upper()}" in upper_href:
                links.append(href)
            if "stat.vin".upper() in upper_href and f"-{vin.upper()}" in upper_href:
                links.append(href)
            if "autohelperbot.com".upper() in upper_href and f"{vin.upper()}_" in upper_href:
                links.append(href)
            if "plc.auction".upper() in upper_href and f"-{vin.upper()}-" in upper_href:
                links.append(href)
            if "plc.ua".upper() in upper_href and f"-{vin.upper()}-" in upper_href:
                links.append(href)
    links = list(set(links))
    [print({"vin": vin, "  link  ": l}) for l in links]
    return [{"vin": vin, "link": l} for l in links]


def images_scrapper(driver, vin):
    links = []
    images_tab = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, images_tab_xpath))
    images_tab.click()
    action = ActionChains(driver)
    sleep(2)
    las = driver.find_elements(By.TAG_NAME, "img")

    for index, la in enumerate(las[10:]):
        try:
            action.move_to_element(la)
            sleep(0.1)
            action.context_click(la).send_keys(Keys.ESCAPE).perform()
            if index % 5 == 0:
                action.send_keys(Keys.ARROW_DOWN).perform()
        except ElementNotInteractableException as e:
            print("No href in tag a")
    las = driver.find_elements(By.TAG_NAME, "a")
    for index, la in enumerate(las):
        href = la.get_attribute("href")
        if href:
            upper_href = href.upper()
            if "bid.cars".upper() in upper_href and f"-{vin.upper()}-" in upper_href:
                links.append(href)
            if "ucars.pro".upper() in upper_href and f"-{vin.upper()}" in upper_href:
                links.append(href)
            if "stat.vin" in upper_href and f"-{vin.upper()}" in upper_href:
                links.append(href)
            if "autohelperbot.com".upper() in upper_href and f"{vin.upper()}_" in upper_href:
                links.append(href)
            if "plc.auction".upper() in upper_href and f"-{vin.upper()}-" in upper_href:
                links.append(href)
            if "plc.ua".upper() in upper_href and f"-{vin.upper()}-" in upper_href:
                links.append(href)

    links = list(set(links))
    [print({"vin": vin, " image_link ": l}) for l in links]
    return [{"vin": vin, "image_link": l} for l in links]

def open_new_request_form(driver):
    driver.get(REMOVE_CONTENT_URL)
    button_new_block_request_element = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, button_new_block_request_xpath))
    button_new_block_request_element.click()
    button_next_block_request_element = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, button_next_block_request_xpath))
    button_next_block_request_element.click()



def block_pages_request(driver, links):
    for index, link in enumerate(links):
        print(index, link["link"], link["vin"])
        open_new_request_form(driver)
        input_block_address = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, input_block_address_xpath))
        input_block_address.send_keys(link["link"])
        # input_block_address.send_keys(link)  # zły link do odtworzenia sytuacji i zabezpieczenia NoSuchElementException
        driver.find_element(By.XPATH, button_send_block).click()
        try:
            print("3 block_pages_request TRY--- OK")
            WebDriverWait(driver, 5).until(lambda d: d.find_element(By.XPATH, request_send_correctly)).click()
            print("3 block_pages_request TRY--- OK - DONE")
            sleep(2)
            continue
        except Exception as e:
            print(f"3 block_pages_request Exception  --- OK ---- bad sign")
        try:
            print("1 block_pages_request TRY - VIN")
            WebDriverWait(driver, 15).until(lambda d: d.find_element(By.XPATH, page_exists_block_address_span))
            driver.find_elements(By.XPATH, "//input")[0].send_keys(link["vin"])
            WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, button_send_block_request)).click()
        except Exception:
            print("1 block_pages_request Exception - VIN")
            pass
        try:
            print("2 block_pages_request TRY - Prześlij prośbę ")
            WebDriverWait(driver, 5).until(lambda d: d.find_element(By.XPATH, button_send_block_request)).click()
            print("2 block_pages_request TRY - Prześlij prośbę -- DDNE")
        except Exception:
            print("2 block_pages_request Exception- Prześlij prośbę ")
            pass



def block_images_request(driver, links):
    # driver.find_element(By.XPATH, button_new_block_request_xpath).click()
        for index, link in enumerate(links):
            print(index, link["image_link"], link["vin"])
            inputs = []
            open_new_request_form(driver)
            try:
                sleep(0.5)
                image_tab_element = WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, image_tab_xpath))
                image_tab_element.click()
            except Exception:
                open_new_request_form(driver)
                try:
                    sleep(0.5)
                    WebDriverWait(driver, 10).until(lambda d: d.find_element(By.XPATH, image_tab_xpath)).click()
                except Exception:
                    for i in range(100):
                        try:
                            locator = f"/html/body/div[{i}]/div[2]/div/div[1]/div/div/div[1]/div/div/span/button[2]/span[3]"
                            driver.find_element(By.XPATH, locator).click()
                            break
                        except Exception:
                            print("--------------------------------------------------------------------------------------")
                            print("----------------------------------------------- TAB GRAFIKA - Exception---------------")
                            print("--------------------------------------------------------------------------------------")
                            continue
            try:
                inputs = driver.find_elements(By.XPATH, "//input")
                inputs[4].click()
            except Exception as e:
                print("Exception  ---------------------   Radio buton w Grafikach nie został znaleziony")
                print(f'{index}.---- {link["image_link"]}  ------   {link["vin"]} --------')
            try:
                inputs[5].click()
                inputs[5].send_keys(link["image_link"])
            except Exception as e:
                print("Exception  ---------------------   Text input w Grafikach nie został znaleziony")
                print(f'{index}. ---- {link["image_link"]}  ------   {link["vin"]} --------')

            driver.find_element(By.XPATH, button_send_block).click()
            try:
                print("1 pop-up  TRY  --- OK ")
                request_send_correctly_button = WebDriverWait(driver, 6).until(lambda d: d.find_element(By.XPATH, request_send_correctly))
                request_send_correctly_button.click()
                print("1 pop-up  TRY  --- OK click")
            except Exception:
                print("1 request_send_correctly Exception  --- OK-  nie wystąpiło")
                pass
            try:
                print("2 button_send_block_request TRY - Prześlij prośbę")
                WebDriverWait(driver, 3).until(lambda d: d.find_element(By.XPATH, button_send_block_request)).click()
            except:
                print("2 button_send_block_request Exception - Prześlij prośbę -  nie wystąpiło")
                pass

            try:
                print("3 passed_address_span TRY - Prześlij prośbę VIN")

                WebDriverWait(driver, 5).until(lambda d: d.find_element(By.XPATH, passed_address_span))
                driver.find_elements(By.XPATH, "//input")[0].send_keys(link["vin"])
                WebDriverWait(driver, 2).until(lambda d: d.find_element(By.XPATH, button_send_block_request)).click()
                print("3 passed_address_span TRY - Prześlij prośbę VIN")
            #     spradzić czy łatwo da się zidentyfikować który pop-up wyskoczył
            except Exception as e:
                print("3 passed_address_span Exception - Prześlij prośbę ")
                try:
                    WebDriverWait(driver, 4).until(lambda d: d.find_element(By.XPATH, button_send_block_request)).click()
                    print("4 button_send_block_request TRY - Prześlij prośbę")
                except Exception:
                    print("4 button_send_block_request Exception - Prześlij prośbę -  nie wystąpiło")
                    pass
                try:
                    request_send_correctly_button = WebDriverWait(driver, 3).until(lambda d: d.find_element(By.XPATH, request_send_correctly))
                    request_send_correctly_button.click()
                    print("5 request_send_correctly TRY --- OK")
                except Exception:
                    print("5 request_send_correctly Exception --- OK")
                    pass
                continue


il = [
]



if __name__ == "__main__":
    import ipdb
    # load_vins()
    driver = init_driver()
    login(driver)
    # first_search(driver)
    page_links = []
    images_links = []

    for vin in VINS:
        new_search(driver, vin)
        # print(f"______________________________________________________vin: {vin}")
        page_links = links_scrapper(driver, vin)
        # print(f"_____________________________________________________________________________page_links       len: {len(page_links)}")
        images_links = images_scrapper(driver, vin)
        # print(f"___________________________________________________________________________images_links       len: {len(images_links)}")
        if page_links:
            block_pages_request(driver, page_links)
        if images_links:
            block_images_request(driver, images_links)
        # import ipdb; ipdb.set_trace()
        # block_images_request(driver, il)



    # print(f"______________________________________________________IMAGES_LINKS len    { len(IMAGES_LINKS) }")
    # try:
    #     block_images_request(driver, IMAGES_LINKS)
    # except Exception:
    #     ipdb.set_trace()

    # print(f"______________________________________________________PAGES_LINKS len    {len(PAGES_LINKS)}")
    # try:
    #
    #     block_pages_request(driver, PAGES_LINKS)
    # except Exception:
    #     ipdb.set_trace()
    ipdb.set_trace()

    driver.quit()
